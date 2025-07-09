import logging
from pathlib import Path
from typing import Dict, List, Any, Generator, Optional

import pandas as pd
import openpyxl

from .base_importer import BaseImporter, ImportResult
from data_importer.core.database import DatabaseManager

logger = logging.getLogger(__name__)


class ExcelImporter(BaseImporter):
    """
    Importer implementation for Excel files (.xlsx).
    Reads data from the first active sheet by default.
    """

    SUPPORTED_EXTENSIONS = [".xlsx"]

    def __init__(self, db_manager: DatabaseManager):
        super().__init__(db_manager)
        logger.info(f"{self.__class__.__name__} initialized.")

    def get_headers(self, file_path: Path) -> List[str]:
        """Reads headers from the first sheet of an Excel file."""
        logger.info(f"Reading headers from Excel file: {file_path}")
        headers = []
        try:
            workbook = openpyxl.load_workbook(file_path, read_only=True)
            sheet = workbook.active

            if sheet.max_row > 0:
                try:  # Try to get the first row cells
                    first_row_cells = sheet[1]  # Get the first row object
                    if first_row_cells:  # Check if the row object is not None
                        for cell in first_row_cells:
                            if cell.value is not None:
                                headers.append(str(cell.value).strip())
                        headers = [
                            h for h in headers if h
                        ]  # Remove empty strings if any
                except IndexError:
                    logger.warning(
                        f"openpyxl could not access row 1 directly for {file_path.name} (sheet might be empty or have no cells in row 1). Treating as no headers by openpyxl."
                    )
                    headers = []  # Ensure headers is empty to trigger pandas or final error

            if not headers:
                logger.info(
                    f"No headers found via openpyxl direct row access for {file_path.name}. Trying pandas fallback."
                )
                try:
                    df_header = pd.read_excel(
                        file_path, sheet_name=0, nrows=0, engine="openpyxl"
                    )
                    headers = [
                        str(h).strip() for h in df_header.columns if str(h).strip()
                    ]
                except pd.errors.EmptyDataError:
                    logger.warning(
                        f"Pandas read_excel found EmptyDataError for {file_path.name} (file is likely empty)."
                    )
                    headers = []
                except Exception as pd_err:
                    logger.warning(
                        f"Pandas fallback for headers failed for {file_path.name}: {pd_err}"
                    )
                    headers = []

            if not headers:
                raise ValueError(
                    f"Excel file '{file_path.name}' appears to have no headers or is empty."
                )

            logger.info(f"Found headers in {file_path.name}: {headers}")
            return headers
        except ValueError as ve:  # Re-raise our specific ValueError
            raise ve
        except Exception as e:  # Catch other unexpected errors
            logger.exception(
                f"General failure to read headers from Excel file {file_path}: {e}"
            )
            # Raise a consistent error message if an unexpected exception occurs
            raise ValueError(
                f"Could not read headers from Excel file {file_path.name} due to an unexpected error: {e}"
            ) from e

    def get_preview(self, file_path: Path, num_rows: int = 5) -> pd.DataFrame:
        """Reads the first N rows of data from the first sheet of an Excel file for preview."""
        logger.info(f"Generating preview ({num_rows} rows) for Excel file: {file_path}")
        try:
            # Pandas is generally robust for previewing Excel
            df = pd.read_excel(
                file_path,
                sheet_name=0,  # First sheet
                nrows=num_rows,
                engine="openpyxl",  # Specify engine
            )
            return df.astype(
                str
            )  # Convert all to string for consistent preview display
        except Exception as e:
            logger.exception(
                f"Failed to generate preview for Excel file {file_path}: {e}"
            )
            # Return empty DataFrame on error
            return pd.DataFrame()

    def read_data(self, file_path: Path) -> Generator[Dict[str, Any], None, None]:
        """
        Reads data row by row from the first sheet of an Excel file.
        Converts each row to a dictionary.
        """
        logger.info(f"Reading data rows from Excel file: {file_path}")
        try:
            workbook = openpyxl.load_workbook(
                file_path, read_only=True, data_only=True
            )  # data_only=True to get values, not formulas
            sheet = workbook.active

            iterator = sheet.rows

            try:
                header_row_values = [
                    str(cell.value).strip() if cell.value is not None else ""
                    for cell in next(iterator)
                ]
            except StopIteration:
                logger.warning(
                    f"Excel file {file_path.name} is empty or has no header row."
                )
                return  # Stop iteration if no header row

            if not any(h for h in header_row_values):  # Check if all headers are empty
                logger.warning(f"Excel file {file_path.name} has an empty header row.")
                # Potentially could try to use pandas to infer headers if this happens
                return

            logger.debug(f"Using headers from Excel: {header_row_values}")

            row_count = 0
            for row in iterator:
                row_count += 1
                values = [cell.value for cell in row]

                # Create a dictionary for the row, handling potential None values and stripping strings
                row_data = {}
                for header, value in zip(header_row_values, values):
                    if header:  # Only include columns with actual headers
                        if isinstance(value, str):
                            row_data[header] = value.strip()
                        elif value is None:
                            row_data[header] = (
                                None  # Or "" if you prefer empty string for None
                            )
                        else:
                            row_data[header] = value

                if not any(row_data.values()):  # Skip entirely empty rows
                    logger.debug(
                        f"Skipping empty row {row_count + 1} in {file_path.name}"
                    )
                    continue

                yield row_data

        except Exception as e:
            logger.exception(
                f"Failed during data reading from Excel file {file_path}: {e}"
            )
            raise ValueError(
                f"Error reading data from Excel file {file_path.name}: {e}"
            ) from e
